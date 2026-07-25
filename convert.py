# -*- coding: utf-8 -*-
"""game_data_v7.xlsx から data_*.js を生成する。
   使い方: python3 tools/convert.py [xlsxパス] [出力先ディレクトリ]"""
import openpyxl, json, sys, os

SRC = sys.argv[1] if len(sys.argv) > 1 else 'game_data_v13.xlsx'
OUT = sys.argv[2] if len(sys.argv) > 2 else '.'
wb = openpyxl.load_workbook(SRC, data_only=True)

def rows(sheet, start):
    for r in wb[sheet].iter_rows(min_row=start, values_only=True):
        if not r[0] or str(r[0]).startswith('※'): continue
        yield r

def js(name, data, header):
    body = json.dumps(data, ensure_ascii=False, indent=2)
    path = os.path.join(OUT, f'{name}.js')
    with open(path, 'w', encoding='utf-8') as f:
        f.write(f'// {header}\n// tools/convert.py により {os.path.basename(SRC)} から自動生成。直接編集しないこと。\n')
        f.write(f'const {name.upper().replace("DATA_","")}_DATA = {body};\n')
    return path

# ── 工程 ──
proc = []
for r in wb['値リスト'].iter_rows(min_row=3, max_row=11, values_only=True):
    if r[6]: proc.append({"id": r[6], "name": r[7], "element": r[8], "effect": r[9]})

# ── アイテム ──
items = []
for r in rows('Items', 4):
    items.append({
        "id": r[0], "name": r[1], "categoryId": r[3],
        "effectId": r[5] or "none", "reactionId": r[7] or None,
        "description": r[8] or "", "note": r[9] or "",
        "source": r[10], "price": r[11], "use": r[12],
        "shopUnlock": r[13], "form": r[14],
    })

# ── レシピ（同名の行をまとめる）──
VESSEL = {"薬瓶": "item_175", "ガラス瓶": "item_088"}
rec, order = {}, []
for r in rows('Recipes', 3):
    name = r[0]
    if name not in rec:
        rec[name] = {"result": name, "vessel": VESSEL.get(r[13]), "ingredients": []}
        order.append(name)
    ing = {"checkType": r[3], "value": r[4], "processId": r[9], "quantity": r[10] or 1}
    if r[6]: ing["checkType2"], ing["value2"] = r[6], r[7]
    if r[11]: ing["note"] = r[11]
    rec[name]["ingredients"].append(ing)
recipes = [rec[n] for n in order]

# ── ヒント ──
hints = []
if 'RecipeHints' in wb.sheetnames:
    for r in rows('RecipeHints', 3):
        hints.append({
            "result": r[0], "no": r[1], "source": r[2],
            "title": r[3], "text": r[4],
            "reliability": r[5], "unlock": r[6] or None,
        })

paths = [
    js('data_processes', proc,  '工程定義'),
    js('data_items',     items, 'アイテム定義'),
    js('data_recipes',   recipes, 'レシピ定義'),
    js('data_hints',     hints, '調合のヒント（紙片）'),
]
print(f'工程 {len(proc)} / アイテム {len(items)} / レシピ {len(recipes)} / ヒント {len(hints)}')
for p in paths: print('  ->', p)
