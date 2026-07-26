# -*- coding: utf-8 -*-
"""game_data_v7.xlsx から data_*.js を生成する。
   使い方: python3 tools/convert.py [xlsxパス] [出力先ディレクトリ]"""
import openpyxl, json, sys, os

SRC = sys.argv[1] if len(sys.argv) > 1 else 'game_data_v25.xlsx'
OUT = sys.argv[2] if len(sys.argv) > 2 else '.'
wb = openpyxl.load_workbook(SRC, data_only=True)

def rows(sheet, start):
    for r in wb[sheet].iter_rows(min_row=start, values_only=True):
        if not r[0] or str(r[0]).startswith('※'): continue
        yield r

def js(name, data, header):
    body = json.dumps(data, ensure_ascii=False, indent=2)
    path = os.path.join(OUT, f'{name}.js')
    with open(path, 'w', encoding='utf-8') as f:
        f.write(f'// {header}\n// tools/convert.py により {os.path.basename(SRC)} から自動生成。直接編集しないこと。\n')
        f.write(f'const {name.upper().replace("DATA_","")}_DATA = {body};\n')
    return path

# ── 工程 ──
proc = []
for r in wb['値リスト'].iter_rows(min_row=3, max_row=11, values_only=True):
    if r[6]: proc.append({"id": r[6], "name": r[7], "element": r[8], "effect": r[9]})

# ── アイテム ──
items = []
for r in rows('Items', 4):
    items.append({
        "id": r[0], "name": r[1], "categoryId": r[3],
        "effectId": r[5] or "none", "reactionId": r[7] or None,
        "description": r[8] or "", "note": r[9] or "",
        "source": r[10], "price": r[11], "use": r[12],
        "shopUnlock": r[13], "form": r[14], "danger": r[15],
    })

# ── レシピ（同名の行をまとめる）──
VESSEL = {"薬瓶": "item_175", "ガラス瓶": "item_088"}
rec, order = {}, []
for r in rows('Recipes', 3):
    name = r[0]
    if name not in rec:
        rec[name] = {"result": name, "vessel": VESSEL.get(r[13]), "ingredients": []}
        order.append(name)
    ing = {"checkType": r[3], "value": r[4], "processId": r[9], "quantity": r[10] or 1}
    if r[6]: ing["checkType2"], ing["value2"] = r[6], r[7]
    if r[11]: ing["note"] = r[11]
    rec[name]["ingredients"].append(ing)
recipes = [rec[n] for n in order]

# ── ヒント ──
hints = []
if 'RecipeHints' in wb.sheetnames:
    for r in rows('RecipeHints', 3):
        hints.append({
            "result": r[0], "no": r[1], "source": r[2],
            "title": r[3], "text": r[4],
            "reliability": r[5],
            "route": r[6], "bundle": r[7], "price": r[8],
        })

# ── 地形・固定マップ ──
terrains, areas = [], []
if 'Terrains' in wb.sheetnames:
    for r in rows('Terrains', 3):
        terrains.append({"id": r[0], "name": r[1], "color": r[2], "description": r[3]})
name2ter = {t["name"]: t["id"] for t in terrains}

if 'Maps' in wb.sheetnames:
    wm = wb['Maps']
    r = 1
    while r <= wm.max_row:
        head = wm.cell(r, 1).value
        if isinstance(head, str) and '（loc_' in head:
            aid = head.split('（')[1].rstrip('）')
            grid = [[name2ter.get(wm.cell(r + 1 + y, 1 + x).value) for x in range(10)] for y in range(10)]
            areas.append({"id": aid, "name": head.split('　')[0].strip(), "grid": grid})
            r += 11
        else:
            r += 1

# ── 採取表 ──
spawns = []
if 'ItemSpawns' in wb.sheetnames:
    for r in rows('ItemSpawns', 3):
        spawns.append({"itemId": r[0], "terrainId": r[1], "timeSlot": r[3],
                       "rate": r[4], "quantity": r[5] or 1, "condition": r[6] or None,
                       "note": r[7] or None})

# ── 旗・出来事・結末 ──
def grab(sheet, keys, start=3):
    out = []
    if sheet not in wb.sheetnames: return out
    for r in wb[sheet].iter_rows(min_row=start, values_only=True):
        if not r[0] or str(r[0]).startswith('※'): continue
        out.append({k: (r[i] if i < len(r) else None) for i, k in enumerate(keys)})
    return out

flags   = grab('Flags',       ['id','name','kind','target','threshold','note'])
events  = grab('Events',      ['id','condition','title','text','effect','target','repeat'])
endings = grab('Endings',     ['id','title','priority','require','forbid','text','altCondition','altText'])
notes   = grab('EndingNotes', ['id','flag','text','forbid'])

# ── 暦 ──
cal = {}
if 'Calendar' in wb.sheetnames:
    wcal = wb['Calendar']
    grab = lambda: [[wcal.cell(r, c).value for c in range(1, 7)] for r in range(1, wcal.max_row + 1)]
    grid, mode = grab(), None
    cal = {"settings": {}, "months": [], "weekdays": [], "moon": [], "weather": {}}
    for row in grid:
        head = row[0]
        if isinstance(head, str) and head.startswith('■'):
            mode = ('settings' if '基本設定' in head
                    else 'moon'     if '月齢' in head
                    else 'weekdays' if '曜日' in head
                    else 'weather'  if '天候' in head
                    else 'months'   if '月'   in head
                    else None)
            continue
        if head in (None, 'key', 'no', 'day', '季節') or (isinstance(head, str) and head.startswith('※')):
            continue
        if mode == 'settings':   cal["settings"][head] = row[1]
        elif mode == 'months':   cal["months"].append({"no": head, "name": row[1], "season": row[2], "note": row[3]})
        elif mode == 'weekdays': cal["weekdays"].append({"no": head, "name": row[1]})
        elif mode == 'moon':     cal["moon"].append({"day": head, "name": row[1], "condition": row[2] or None})
        elif mode == 'weather':  cal["weather"][head] = {"晴": row[1], "曇": row[2], "雨": row[3], "霧": row[4], "雪": row[5]}

paths = [
    js('data_processes', proc,  '工程定義'),
    js('data_items',     items, 'アイテム定義'),
    js('data_recipes',   recipes, 'レシピ定義'),
    js('data_hints',     hints, '調合のヒント（紙片）'),
    js('data_calendar',  cal, '暦・月齢'),
    js('data_terrains',  {"terrains": terrains, "areas": areas}, '地形と固定マップ'),
    js('data_spawns',    spawns, '採取表（地形×時間帯）'),
    js('data_flags',     flags,   '旗'),
    js('data_events',    events,  '出来事'),
    js('data_endings',   {"endings": endings, "notes": notes}, '結末と後日談の断片'),
]
print(f'工程 {len(proc)} / アイテム {len(items)} / レシピ {len(recipes)} / ヒント {len(hints)} / 暦 {len(cal.get("months", []))}ヶ月 / 地形 {len(terrains)} / マップ {len(areas)} / 採取 {len(spawns)} / 旗 {len(flags)} / 出来事 {len(events)} / 結末 {len(endings)}')
for p in paths: print('  ->', p)
